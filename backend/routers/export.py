import csv
import io
import json
import os
import sqlite3
import tempfile
from datetime import datetime, timezone

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query
from fastapi.responses import FileResponse, Response
from sqlalchemy.orm import Session

from database import get_db
from auth import require_user_id
from services.sync_to_root import sync_all_to_root
from models import (
    Author, Book, Publisher, Brand, BookSeries, Category,
    Bookshelf, BookCollection,
)

router = APIRouter(prefix="/api/export", tags=["export"])

SCOPES = {
    "books":        (Book,           "books"),
    "authors":      (Author,         "authors"),
    "publishers":   (Publisher,      "publishers"),
    "brands":       (Brand,          "brands"),
    "series":       (BookSeries,     "book_series"),
    "categories":   (Category,       "categories"),
    "bookshelves":  (Bookshelf,      "bookshelves"),
    "collections":  (BookCollection, "book_collections"),
}

FORMATS = ("sql", "csv", "excel", "markdown", "json")

EXTENSIONS = {
    "sql":      ".sql",
    "csv":      ".csv",
    "excel":    ".xlsx",
    "markdown": ".md",
    "json":     ".json",
}

CONTENT_TYPES = {
    "sql":      "application/sql",
    "csv":      "text/csv",
    "excel":    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "markdown": "text/markdown",
    "json":     "application/json",
}


def _columns(model):
    return [(c.name, str(c.type)) for c in model.__table__.columns]


def _rows(db, model):
    col_names = [c[0] for c in _columns(model)]
    result = []
    for obj in db.query(model).all():
        row = {}
        for name in col_names:
            val = getattr(obj, name, None)
            if hasattr(val, "isoformat"):
                val = val.isoformat()
            row[name] = val
        result.append(row)
    return col_names, result


def _build_json(col_names, data):
    return json.dumps(data, ensure_ascii=False, indent=2, default=str)


def _build_csv(col_names, data):
    out = io.StringIO()
    w = csv.DictWriter(out, fieldnames=col_names)
    w.writeheader()
    w.writerows(data)
    return out.getvalue()


def _md_cell(val):
    """Format a value for a Markdown table cell, escaping pipes and newlines."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "True" if val else "False"
    s = str(val).replace("|", "\\|").replace("\n", " ").replace("\r", "")
    return s


def _build_markdown(col_names, data):
    lines = []
    lines.append("| " + " | ".join(col_names) + " |")
    lines.append("| " + " | ".join(["---"] * len(col_names)) + " |")
    for row in data:
        vals = [_md_cell(row.get(c)) for c in col_names]
        lines.append("| " + " | ".join(vals) + " |")
    return "\n".join(lines)


def _build_sql(col_names, data, table_name):
    lines = [f"-- Export of {table_name}  ({len(data)} rows)", ""]
    for row in data:
        vals = []
        for c in col_names:
            v = row.get(c)
            if v is None:
                vals.append("NULL")
            elif isinstance(v, bool):
                vals.append("1" if v else "0")
            elif isinstance(v, (int, float)):
                vals.append(str(v))
            else:
                escaped = str(v).replace("'", "''").replace("\n", "\\n").replace("\r", "")
                vals.append(f"'{escaped}'")
        lines.append(
            f"INSERT INTO {table_name} ({', '.join(col_names)}) "
            f"VALUES ({', '.join(vals)});"
        )
    return "\n".join(lines)


def _build_excel(col_names, data):
    import openpyxl
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "export"
    for i, name in enumerate(col_names, 1):
        ws.cell(row=1, column=i, value=name)
    for r, row in enumerate(data, 2):
        for c, name in enumerate(col_names, 1):
            ws.cell(row=r, column=c, value=row.get(name))
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


BUILDERS = {
    "json":     _build_json,
    "csv":      _build_csv,
    "markdown": _build_markdown,
    "sql":      _build_sql,
    "excel":    _build_excel,
}


def _snapshot_db(db_path: str) -> str:
    """Copy the SQLite file to a temp path via the backup API.

    The backup API produces a consistent, self-contained snapshot even if the
    live database has concurrent connections or uncommitted transactions."""
    fd, tmp = tempfile.mkstemp(prefix="my-library-export-", suffix=".db")
    os.close(fd)
    src = dst = None
    try:
        src = sqlite3.connect(db_path)
        dst = sqlite3.connect(tmp)
        src.backup(dst)
    except BaseException:
        os.unlink(tmp)
        raise
    finally:
        if dst is not None:
            dst.close()
        if src is not None:
            src.close()
    return tmp


@router.get("/database")
def export_database(
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    user_id: int = Depends(require_user_id),
):
    """Download the current user's entire database as a SQLite file."""
    bind = db.get_bind()
    engine = getattr(bind, "engine", bind)
    db_path = engine.url.database
    if not db_path or not os.path.isfile(db_path):
        raise HTTPException(status_code=404, detail="Database file not found")

    snapshot = _snapshot_db(db_path)
    background_tasks.add_task(os.unlink, snapshot)

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    return FileResponse(
        snapshot,
        media_type="application/vnd.sqlite3",
        filename=f"my-library-database-{stamp}.db",
    )


@router.get("/")
def export_data(
    fmt: str = Query(..., alias="format", description=f"Export format: {', '.join(FORMATS)}"),
    scope: str = Query(..., description=f"Export scope: {', '.join(SCOPES.keys())}"),
    db: Session = Depends(get_db),
    user_id: int = Depends(require_user_id),
):
    if scope not in SCOPES:
        raise HTTPException(status_code=400, detail=f"Invalid scope. Choose from: {', '.join(SCOPES.keys())}")
    if fmt not in FORMATS:
        raise HTTPException(status_code=400, detail=f"Invalid format. Choose from: {', '.join(FORMATS)}")

    model, table_name = SCOPES[scope]
    col_names, data = _rows(db, model)

    builder = BUILDERS[fmt]
    kwargs = {}
    if fmt == "sql":
        kwargs["table_name"] = table_name
    content = builder(col_names, data, **kwargs)

    ext = EXTENSIONS[fmt]
    content_type = CONTENT_TYPES[fmt]
    return Response(
        content=content,
        media_type=content_type,
        headers={"Content-Disposition": f'attachment; filename="{scope}{ext}"'},
    )


@router.post("/sync-to-root")
def sync_to_root(
    differential: bool = Query(True, description="Only sync changed/missing entries"),
    db: Session = Depends(get_db),
    user_id: int = Depends(require_user_id),
):
    """Sync all entities from the current user's database to root.db.

    Set differential=false for a full resync of every row."""
    counts = sync_all_to_root(db, differential=differential)
    return {"message": "Sync complete", "differential": differential, "counts": counts}
